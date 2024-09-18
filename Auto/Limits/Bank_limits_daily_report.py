#!/usr/bin/env python
# coding: utf-8

# In[13]:


print('The "Limits" just started running')


# In[14]:


import pandas as pd
import numpy as np
from datetime import date, timedelta
import seaborn as sns
import matplotlib.pyplot as plt
import matplotlib.axes as ax
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.drawing.image import Image
import win32com.client as win32
import os
from PIL import ImageGrab
import win32com.client

olApp = win32.Dispatch('Outlook.Application')
olNS = olApp.GetNameSpace('MAPI')

import sys
try:
    sys.path.append("C:\\Users\\KlimovaAnnaA\\Documents\\MyFiles\\Projects\\Working_attributes")
except:
    sys.path.append("z:\\Anna_Klimova\\Working_attributes")
from Imports import * 


# In[15]:


run_from_Z = False # True/False Заполните это поле, если работа отчета производится с общего диска

Print_to_excel = True # Создать excel-файл с расчетами? True/False
Display_mail = True # Показать письмо для отправки? True/False
Send_mail = True # Создать и отправить письмо с расчетами и графиком? True/False

mail_to_Ech = 'TarakanovMIu@aimmngt.com;Aleksandr.A.Ovchinnikov@eurochem.ru;Anastasiya.Barmenkova@eurochem.ru;denis.perevezentsev@rowanfm.cn;Tatyana.Votyakova@eurochem.ru;Aleksandr.Ostreyko@eurochem.ru;IvashovaEA@rowanfm.com;Axana.Davletkireeva@eurochemgroup.ae; ZubkovNA@aimmngt.com; <Pavel.Burundukov@rowanfm.ae>; GerchakAI@aimmngt.com;TimokhinEV@aimmngt.com;Pavel.Lyubin@greenms.ae;KotinaOI@aimmngt.com;naumovaei@rowanfm.com' # Получатель письма
mail_ECH_SAM ='<TarakanovMIu@aimmngt.com>; denis.perevezentsev@rowanfm.cn; <Aleksandr.Ostreyko@eurochem.ru>;<persio.ravena@eurochemsam.com;<helio.pimentel@eurochemsam.com>; <renato.costa@eurochemsam.com>; <joao.cruz@eurochemsam.com>; <lua.pereira@eurochemsam.com>;  TimokhinEV@aimmngt.com; Pavel.Lyubin@greenms.ae;'
mail_ECH_NAM = '<TarakanovMIu@aimmngt.com>; denis.perevezentsev@rowanfm.cn; <Aleksandr.Ostreyko@eurochem.ru>;  Donal.Lambert@eurochem-na.com;<Denis.Bukin@eurochem-na.com>; TimokhinEV@aimmngt.com; Pavel.Lyubin@greenms.ae;'
mail_ECH_Europe ='<TarakanovMIu@aimmngt.com>; denis.perevezentsev@rowanfm.cn; <Aleksandr.Ostreyko@eurochem.ru>; <Aleksandr.Vasilyev@eurochemgroup.com>; TimokhinEV@aimmngt.com; Pavel.Lyubin@greenms.ae;'

mail_to_Suek = 'TarakanovMIu@aimmngt.com;Aleksandr.A.Ovchinnikov@eurochem.ru;Anastasiya.Barmenkova@eurochem.ru;denis.perevezentsev@rowanfm.cn;MaltsevaAA@rowanfm.com;ChernicherBA@suek.ru;IvashovaEA@rowanfm.com;ZubkovNA@aimmngt.com; <Pavel.Burundukov@rowanfm.ae>; ruslan.minikeev@black-sand-commodities.ae;TsuranAS@suek.ru;PanteleevaTaV@suek.ru;GerchakAI@aimmngt.com;TimokhinEV@aimmngt.com;Pavel.Lyubin@greenms.ae;KotinaOI@aimmngt.com;naumovaei@rowanfm.com' # Получатель письма
mail_from = 'KlimovaAnnaA@aimmngt.com'
# signature = """Maksim Tarakanov <br><br>
    # Whatsapp: +7 915 161 29 12<br>
    # Financial risk management"""
signature = "Financial risk management"
# mail_from = 'TarakanovMIu@aimmngt.com' # mail from


# In[16]:


mail_ECH_SAM, mail_ECH_NAM, mail_ECH_Europe


# In[17]:


query = f"""
SELECT *
FROM "RISKACCESS"."bankAccountsBalanceDaily"
WHERE "reportDate" = (SELECT MAX("reportDate") FROM "RISKACCESS"."bankAccountsBalanceDaily")
"""
bankAccountsBalanceDaily_data = export_from_RISKCUSTOM(query) # export data
BABD_data_work = bankAccountsBalanceDaily_data # BABD_data_work,
BABD_data_work['balanceUsd_mln'] = BABD_data_work.balanceUsd/10**6 # balanceUsd_mln
# balanceUsd_activmoney_market
accountStatus_list = ['active', 'mmarket']
BABD_data_work['balanceUsd_activmoney_market'] = 0
BABD_data_work.loc[BABD_data_work.accountStatus.isin(accountStatus_list), 'balanceUsd_activmoney_market'] = BABD_data_work.loc[BABD_data_work.accountStatus.isin(accountStatus_list), 'balanceUsd_mln']

query = f"""
SELECT CONCAT(CONCAT("bankId", "holding"), "limitType") as "con", "activeFrom", "bankId", "limitType", "limit", "holding"
FROM "RISKACCESS"."xxmrBankLimits" 
WHERE "limitType" IN ('Transit', 'Deposit')
"""
Limit_export_data = export_from_RISKCUSTOM(query) # export data
LE_data_work = Limit_export_data.sort_values(by='activeFrom').drop_duplicates(subset='con', keep='last')  # Unique strings with lasr value for activeFrom
# pivot by bankid and mln of limit
LE_data_work['H_B_concat'] = LE_data_work.holding + '_' + LE_data_work.bankId
LE_data_group = pd.pivot_table(data=LE_data_work, index=['H_B_concat', 'bankId', 'holding'], values='limit', aggfunc='sum').reset_index()
LE_data_group.limit = LE_data_group.limit/10**6
# merge BABD_data_work and LE_data_group
BABD_data_work = BABD_data_work.reset_index(drop=True)
BABD_data_work['H_B_concat'] = BABD_data_work.holding + '_' + BABD_data_work.bankId
BABD_data_work['Limit'] = BABD_data_work.merge(LE_data_group, how='left', left_on='H_B_concat', right_on='H_B_concat', validate='many_to_one').iloc[:,-1]
# Usage and Usage_activmoney_market
BABD_data_work['Usage_activmoney_market'] = 0
BABD_data_work['Usage'] = 0
BABD_data_limit_not_na = BABD_data_work[~BABD_data_work.Limit.isna()]
BABD_data_limit_not_na['Usage_activmoney_market'] = (BABD_data_limit_not_na.balanceUsd_activmoney_market/BABD_data_limit_not_na.Limit)*100
BABD_data_limit_not_na['Usage'] = (BABD_data_limit_not_na.balanceUsd_mln/BABD_data_limit_not_na.Limit)*100
BABD_data_work[~BABD_data_work.Limit.isna()] = BABD_data_limit_not_na
BABD_data_work.loc[(BABD_data_work.Usage == np.inf) | (BABD_data_work.Usage.isna()) | (BABD_data_work.Usage == -np.inf), 'Usage'] = 0
BABD_data_work.loc[(BABD_data_work.Usage_activmoney_market == np.inf) | (BABD_data_work.Usage_activmoney_market.isna()) | (BABD_data_work.Usage_activmoney_market == -np.inf), 'Usage_activmoney_market'] = 0
BABD_data_work['Segment'] = merge_SalesUnits(df=BABD_data_work, merge_col='businessSegmentDetailed', col='buCode') # merge Segment
# Bank_name
query = """SELECT "bankId", "name", "country"
FROM "RISKACCESS"."xxmrBankLimitsBanks"
"""
data_xxmrBankLimitsBanks = export_from_RISKCUSTOM(query)
data_BLB = data_xxmrBankLimitsBanks.drop_duplicates(subset='bankId').dropna()
BABD_data_work['bank_name'] = BABD_data_work.merge(data_BLB, how='left', left_on='bankId', right_on='bankId', validate='many_to_one').iloc[:, -2]
BABD_data_work.loc[(BABD_data_work.bank_name.isna()), 'bank_name'] = BABD_data_work.loc[(BABD_data_work.bank_name.isna()), 'bankName']
# rounding
BABD_data_work[['Usage_activmoney_market','Usage']] = BABD_data_work[['Usage_activmoney_market','Usage']].apply(lambda x:round(x, 1))
BABD_data_work[['balanceUsd_mln','balanceUsd_activmoney_market', 'Limit']] = BABD_data_work[['balanceUsd_mln','balanceUsd_activmoney_market', 'Limit']].round(2)

### TO EXCEL
BABD_data_work = BABD_data_work.rename(columns={'balanceUsd_activmoney_market':'Active', 'balanceUsd_mln':'Total', 'Usage_activmoney_market':'%_active', 'Usage':'%_total'})
BABD_data_work = BABD_data_work[(BABD_data_work.Active!=0)&(BABD_data_work.Limit!=0)]
# Creating Excel Writer Object from Pandas 
report_date = str(BABD_data_work.reportDate.max())[:10]
# by holding
holding_list = BABD_data_work.holding.unique().tolist()
for holding in holding_list:
    holding_data = BABD_data_work[BABD_data_work.holding == holding].reset_index(drop=True) # holding data
    holding_data['Total_bank_name'] = holding_data.groupby('bank_name')['Total'].transform('sum')
    # table 1
    tabel_bankName = pd.pivot_table(data=holding_data, 
                    index='bank_name', 
                    values=['Limit', 'Active', 'Total', '%_active', '%_total'], 
                    aggfunc={'Limit':'mean', 'Active':'sum', 'Total':'sum', '%_active':'sum', '%_total':'sum'},
                    fill_value=0)\
                    .reset_index()\
                    .sort_values(['Active', 'Total'], ascending=False)
    # tabel_bankName = tabel_bankName[['bank_name','Limit', 'Active', '%_active', 'Total', '%_total', 'Segment']]
    tabel_bankName = tabel_bankName[['bank_name','Limit', 'Active', '%_active', 'Total', '%_total']]
    # table 2
    table_bankCountryCode = pd.pivot_table(data=holding_data, 
                    index='bankCountryCode', 
                    values=['Active', 'Total'], 
                    aggfunc={'Active':'sum', 'Total':'sum'})\
                    .reset_index()\
                    .sort_values(['Active', 'Total'], ascending=False)
    # table 3
    table_Segment = pd.pivot_table(data=holding_data, 
                    index='Segment', 
                    values=['Active', 'Total'], 
                    aggfunc={'Active':'sum', 'Total':'sum'})\
                    .reset_index()\
                    .sort_values(['Active', 'Total'], ascending=False)
    # sheet 3
    # Create the frame in the right order 
    tabel_bankName_2 = pd.pivot_table(data=holding_data, 
                    index=['bank_name','bankCountryCode'], 
                    values=['Active', 'Total'], 
                    aggfunc={'Active':'sum', 'Total':'sum'},
                    fill_value=0)\
                    .reset_index()\
                    .sort_values(['Active', 'Total'], ascending=False)
    Frame_table = table_bankCountryCode.merge(tabel_bankName_2, how='left', left_on='bankCountryCode', right_on='bankCountryCode').dropna()
    # create columns
    new_columns = table_Segment.Segment.values.tolist()
    Frame_table = Frame_table.reindex(columns=Frame_table.columns.tolist() + new_columns, fill_value=0)
    Frame_table = Frame_table.drop(['Active_x','Total_x'], axis=1).rename(columns={'Active_y':'Active','Total_y':'Total'})
    # fill values
    pivot_data_for_frame = holding_data.pivot_table(index='bank_name', values='Total', aggfunc=sum, columns='Segment').reset_index()
    df_dict_list = pivot_data_for_frame.to_dict(orient='records')
    for bank_name in Frame_table.bank_name:
        for segment_column in Frame_table.iloc[:, 4:].columns:
            Frame_table.loc[Frame_table.bank_name==bank_name,segment_column] = [df_dict_list[x] for x in range(len(df_dict_list)) if df_dict_list[x]['bank_name']==bank_name][0][segment_column]
    # filter banks with empty Active and Total
    empty_banks_list = tabel_bankName[(tabel_bankName.Active==0)&(tabel_bankName.Total==0)].bank_name.unique().tolist()
    Frame_table = Frame_table[~Frame_table.bank_name.isin(empty_banks_list)].fillna(0)
    ### add final country rows
    Frame_table = Frame_table.reset_index(drop=True)
    table_bankCountryCode_for_concat = table_bankCountryCode[table_bankCountryCode.bankCountryCode.isin(Frame_table.bankCountryCode.unique().tolist())].reset_index(drop=True)
    data_concat = pd.concat([Frame_table,table_bankCountryCode_for_concat])
    # work with index
    data_concat['index_copy'] = data_concat.index
    change_values_index_dict = {}
    for table_index in range(1,Frame_table.shape[0]):
        if Frame_table.loc[table_index, 'bankCountryCode']!=Frame_table.loc[(table_index-1), 'bankCountryCode']:
            change_values_index_dict[Frame_table.loc[(table_index-1), 'bankCountryCode']] = table_index-0.5
    change_values_index_dict[Frame_table.loc[len(Frame_table)-1, 'bankCountryCode']] = len(Frame_table)-0.5
    # change index
    for country in list(change_values_index_dict.keys()):
        data_concat.loc[(data_concat.bankCountryCode==country)&(data_concat.bank_name.isna()),'index_copy'] = change_values_index_dict[country]
    Frame_table = data_concat.set_index('index_copy').sort_index()
    # fill values
    pivot_data_for_frame = holding_data.pivot_table(index='bankCountryCode', values='Total', aggfunc=sum, columns='Segment').reset_index()
    df_dict_list = pivot_data_for_frame.to_dict(orient='records')
    for bankCountryCode in Frame_table.bankCountryCode.unique().tolist():
        for segment_column in Frame_table.iloc[:, 4:].columns:
            Frame_table.loc[(Frame_table.bankCountryCode==bankCountryCode)&(Frame_table.bank_name.isna()),segment_column] = [df_dict_list[x] for x in range(len(df_dict_list)) if df_dict_list[x]['bankCountryCode']==bankCountryCode][0][segment_column]

    
    # tabel_bankName['Segment'] = tabel_bankName.merge(holding_data, how='left', left_on='bank_name', right_on='bank_name')['Segment']
    # table_bankCountryCode['Segment'] = table_bankCountryCode.merge(holding_data, how='left', left_on='', right_on='')

    if holding == 'SUEK':
        tabel_bankName_SUEK = tabel_bankName
        table_bankCountryCode_SUEK = table_bankCountryCode
        table_Segment_SUEK = table_Segment
        Frame_table_SUEK = Frame_table
        change_values_index_dict_SUEK = change_values_index_dict
    else:
        tabel_bankName_Ech = tabel_bankName
        tabel_bankName_Ech = tabel_bankName_Ech[~tabel_bankName_Ech.bank_name.str.contains('Urbo')] # фильтр на конкретный банк
        table_bankCountryCode_Ech = table_bankCountryCode
        table_Segment_Ech = table_Segment
        Frame_table_Ech = Frame_table
        change_values_index_dict_Ech = change_values_index_dict
    # to excel
    if Print_to_excel == True:
        Output_file = '_'.join([report_date, holding,'limits_report.xlsx'])
        writer = pd.ExcelWriter(Output_file, engine='openpyxl')  
        workbook=writer.book
        pd.DataFrame({'holding':f'{holding} (in MUSD)'}, index=[1]).to_excel(writer, sheet_name=holding, index=False, header=False)
        tabel_bankName.to_excel(writer, sheet_name=holding, index=False, startrow=1)
        table_bankCountryCode.to_excel(writer, sheet_name=holding, startcol=8, index=False, startrow=1)
        table_Segment.to_excel(writer, sheet_name=holding, startcol=13, index=False, startrow=1)
        writer.close()

        # writing new files for segments
        if holding=='EUROCHEM':
            sheets_list_Ech = ['EUROCHEM']
            sheets_dict_Ech = {'EUROCHEM':Output_file}
            for filt in [['SAM'], ['NAM'], ['EUROPE', 'EUROPE distributors and plants', 'SUEK AG+']]:
                sheet_n = f'{holding}_{filt[0]}'
                sheets_list_Ech = sheets_list_Ech + [sheet_n]

                Output_file_filt = '_'.join([report_date, holding, filt[0],'limits_report.xlsx'])
                sheets_dict_Ech[f'EUROCHEM_{filt[0]}'] = Output_file_filt
                writer = pd.ExcelWriter(Output_file_filt, engine='openpyxl')  
                workbook=writer.book

                pd.DataFrame({'holding':f'{holding} (in MUSD)'}, index=[1]).to_excel(writer, sheet_name=sheet_n, index=False, header=False)
                
                holding_data.Segment = holding_data.Segment.replace({'TRADING Europe':'EUROPE'})
                tabel_bankName_filt = holding_data[holding_data.Segment.isin(filt)]
                tabel_bankName_filt_pivot = pd.pivot_table(data=tabel_bankName_filt, index='bank_name', 
                                                values=['Limit', 'Active', 'Total_bank_name'], 
                                                aggfunc={'Limit':'mean', 'Active':'sum', 'Total_bank_name':'mean'},
                                                fill_value=0)\
                                                .reset_index()\
                                                .sort_values(['Active', 'Total_bank_name'], ascending=False)
                tabel_bankName_filt_pivot['%_active'] = (tabel_bankName_filt_pivot.Active/tabel_bankName_filt_pivot.Limit)*100
                tabel_bankName_filt_pivot['%_total'] = (tabel_bankName_filt_pivot.Total_bank_name/tabel_bankName_filt_pivot.Limit)*100
                tabel_bankName_filt_pivot[['%_total', '%_active']] = tabel_bankName_filt_pivot[['%_total', '%_active']].fillna(0)
                tabel_bankName_filt_pivot.loc[tabel_bankName_filt_pivot['%_active']==np.inf, '%_active'] = 0
                tabel_bankName_filt_pivot.loc[tabel_bankName_filt_pivot['%_total']==np.inf, '%_total'] = 0
                tabel_bankName_filt_pivot = tabel_bankName_filt_pivot.rename(columns={'Total_bank_name':'Total'})
                tabel_bankName_filt_pivot = tabel_bankName_filt_pivot[['bank_name','Limit', 'Active', '%_active', 'Total', '%_total']]
                
                tabel_bankName_filt_pivot.to_excel(writer, sheet_name=sheet_n, index=False, startrow=1)
                # table_bankCountryCode.to_excel(writer, sheet_name=sheet_n, startcol=8, index=False, startrow=1)
                # table_Segment.to_excel(writer, sheet_name=sheet_n, startcol=13, index=False, startrow=1)
                writer.close()

        new_list(bankAccountsBalanceDaily_data[bankAccountsBalanceDaily_data.holding==holding], sheet_name='data', output_file=Output_file)
        new_list(Frame_table, sheet_name='Banks_to_segments_(Total)', output_file=Output_file)
        if holding == 'SUEK':
            Output_file_SUEK = Output_file
        else:
            Output_file_Ech = Output_file


# In[18]:


BABD_data_work.loc[BABD_data_work.Segment=='External',['holding','buCode','buName','Segment']]
BABD_data_work.loc[:,'Segment'].unique()


# In[19]:


### FORMAT
if Print_to_excel == True:
    SUEK_tables_list = [tabel_bankName_SUEK, table_bankCountryCode_SUEK, table_Segment_SUEK, Frame_table_SUEK]
    Ech_tabels_list = [tabel_bankName_Ech, table_bankCountryCode_Ech, table_Segment_Ech, Frame_table_Ech]
            
    holdind = ''
    Output_file = ''
    len_A_dict = {}
    for holding in holding_list:
        if holding == 'SUEK':
            tables_list = SUEK_tables_list
            Output_file = Output_file_SUEK
            sheet_list = ['SUEK']
            sheet_dict = {'SUEK': Output_file_SUEK}
        else:
            tables_list = Ech_tabels_list
            Output_file = Output_file_Ech
            sheet_list = sheets_list_Ech
            sheet_dict = sheets_dict_Ech

        # open file
        for sheet_n in sheet_list:
            Output_file = f'{sheet_dict[sheet_n]}'
            wb = openpyxl.load_workbook(Output_file)
            ws = wb[sheet_n]
            
            len_A=0
            for row in range(3,10000):
                if type(ws[f'A{row}'].value) is not str:
                    len_A = row-1
                    break
            len_A_dict[sheet_n] = len_A
            data = []
            for row in ws[f'A2:F{len_A}']:
                row_data = [cell.value for cell in row]
                data.append(row_data)
            data = pd.DataFrame(data)
            data.columns = data.iloc[0,:]
            tables_list[0] = data.iloc[1:,:]

            # color
            color_areas_list = [f"C2:C{len(tables_list[0])+2}", f"J2:J{len(tables_list[1])+2}", f"O2:O{len(tables_list[2])+2}"]
            cell_color = PatternFill(start_color='00FFCC99', end_color='00FFCC99', fill_type = "solid")
            for color_area in color_areas_list:
                for row in ws[color_area]:
                    for cell in row:
                        cell.fill = cell_color
                if sheet_n!=holding:
                    break
            cell_color = PatternFill(start_color='00CCFFCC', end_color='00CCFFCC', fill_type = "solid")
            for row in ws['A1:F1']:
                for cell in row:
                    cell.fill = cell_color
                    cell.font = Font(bold=True)
            # Borders         
            medium = Side(border_style="medium", color="000000")
            right_line_areas_list = [f"F2:F{len(tables_list[0])+2}", f"K2:K{len(tables_list[1])+2}", f"P2:P{len(tables_list[2])+2}"]
            for right_line_area in right_line_areas_list:
                for row in ws[right_line_area]:
                    for cell in row:
                        cell.border = Border(top=None, left=None, right=medium, bottom=None)
                if sheet_n!=holding:
                    break
            left_line_areas_list = [f"A2:A{len(tables_list[0])+2}", f"I2:I{len(tables_list[1])+2}", f"N2:N{len(tables_list[2])+2}"]
            for left_line_area in left_line_areas_list:
                for row in ws[left_line_area]:
                    for cell in row:
                        cell.border = Border(top=None, left=medium, right=medium, bottom=None)
                if sheet_n!=holding:
                    break                        
            top_line_areas_list = ['A2:F2', 'I2:K2', 'N2:P2']
            for top_line_area in top_line_areas_list:
                for row in ws[top_line_area]:
                    for cell in row:
                        cell.border = Border(top=medium, left=medium, right=medium, bottom=medium)
                if sheet_n!=holding:
                    break
            bottom_line_areas_list = [f'A{len(tables_list[0])+2}:F{len(tables_list[0])+2}', f'I{len(tables_list[1])+2}:K{len(tables_list[1])+2}', f'N{len(tables_list[2])+2}:P{len(tables_list[2])+2}']
            for bottom_line_area in bottom_line_areas_list:
                for row in ws[bottom_line_area]:
                    for cell in row:
                        cell.border = Border(top=None, bottom=medium)
                if sheet_n!=holding:
                    break
            # font color errors
            tables_list[0] = tables_list[0].reset_index(drop=True)
            # tables with deviations
            significant_deviation = tables_list[0][(((tables_list[0].Active - tables_list[0].Limit) > tables_list[0].Limit*0.1) | ((tables_list[0].Active - tables_list[0].Limit) > 10)) & ((tables_list[0].Active - tables_list[0].Limit) > 0)]
            significant_deviation_index_liist = significant_deviation.index.tolist()
            insignificant_deviation = tables_list[0][(((tables_list[0].Active - tables_list[0].Limit) <= tables_list[0].Limit*0.1) & ((tables_list[0].Active - tables_list[0].Limit) <= 10)) & ((tables_list[0].Active - tables_list[0].Limit) > 0)]
            insignificant_deviation_index_liist = insignificant_deviation.index.tolist()
            # areas lists
            significant_deviation_areas_liist = [f'A{x+3}:F{x+3}' for x in significant_deviation_index_liist] # red
            insignificant_deviation_areas_liist = [f'A{x+3}:F{x+3}' for x in insignificant_deviation_index_liist] # orange
            # color areas
            cell_color = PatternFill(start_color='00FF8080', end_color='00FF8080', fill_type = "solid")
            for color_area in significant_deviation_areas_liist:
                for row in ws[color_area]:
                    for cell in row:
                        cell.fill = cell_color # red
            for color_area in insignificant_deviation_areas_liist:
                for row in ws[color_area]:
                    for cell in row:
                        cell.font = Font(color="00FF9900") # orange
            # Weight of olumns
            ws.column_dimensions['A'].width = 29
            ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 8
            ws.column_dimensions['E'].width = 8
            ws.column_dimensions['F'].width = 8
            ws.column_dimensions['G'].width = 5
            ws.column_dimensions['H'].width = 5
            ws.column_dimensions['I'].width = 16
            ws.column_dimensions['J'].width = 8
            ws.column_dimensions['K'].width = 8
            ws.column_dimensions['L'].width = 5
            ws.column_dimensions['M'].width = 5
            ws.column_dimensions['N'].width = 28
            ws.column_dimensions['O'].width = 8
            ws.column_dimensions['P'].width = 8
            # Rounding
            rounding_2_numbers_list = [f'C3:C{len(tables_list[0])+2}', f'E3:E{len(tables_list[0])+2}', f'J3:K{len(tables_list[1])+2}', f'O3:P{len(tables_list[2])+2}']
            rounding_1_numbers_list = [f'D3:D{len(tables_list[0])+2}', f'F3:F{len(tables_list[0])+2}']
            for color_area in rounding_2_numbers_list:
                for row in ws[color_area]:
                    for cell in row:
                        cell.number_format = '0.00'
            for color_area in rounding_1_numbers_list:
                for row in ws[color_area]:
                    for cell in row:
                        cell.number_format = '0.0'
            # close file
            wb.save(Output_file)
            wb.close() 
            
            if sheet_n==holding:
                # Format 3-th sheet
                # open file
                wb = openpyxl.load_workbook(Output_file)
                ws = wb['Banks_to_segments_(Total)']
                # Frame table line after countries
                # Number of strings to line
                change_values_index_list = []
                tables_list[3] = tables_list[3].reset_index(drop=True)
                for table_index in range(1,tables_list[3].shape[0]):
                    if tables_list[3].loc[table_index, 'bankCountryCode']!=tables_list[3].loc[(table_index-1), 'bankCountryCode']:
                        change_values_index_list += [table_index+1]
                change_values_index_list += [len(tables_list[3])+1]
                # Letter of the end of the table
                letters_dict = {i:chr(i+64) for i in range(1,27)}
                letter_last_column = letters_dict[tables_list[3].shape[1]]
                # right lines 
                lines_areas = [f'{i}1:{i}{tables_list[3].shape[0]+1}' for i in ['B','D',letter_last_column]]
                for bottom_line_area in lines_areas:
                    for row in ws[bottom_line_area]:
                        for cell in row:
                            cell.border = Border(top=cell.border.top, left=cell.border.left, right=medium, bottom=cell.border.bottom)
                # bottom lines
                lines_areas = [f'A{i}:{letter_last_column}{i}' for i in change_values_index_list]+[f'A1:{letter_last_column}1']
                for bottom_line_area in lines_areas:
                    for row in ws[bottom_line_area]:
                        for cell in row:
                            cell.border = Border(top=cell.border.top, left=cell.border.left, right=cell.border.right, bottom=medium)
                # color of total by country 
                color_areas = [f'A{i}:{letter_last_column}{i}' for i in change_values_index_list]
                # close file
                cell_color = PatternFill(start_color='00CCFFCC', end_color='00CCFFCC', fill_type = "solid")
                for color_area in color_areas:
                    for row in ws[color_area]:
                        for cell in row:
                                cell.fill = cell_color
                                cell.font = Font(bold=True)
                # Rounding
                for row in ws[f'C2:{letter_last_column}{len(tables_list[3])+1}']:
                    for cell in row:
                        cell.number_format = '0.00'
                # Weight of olumns
                ws.column_dimensions['B'].width = 29
        
                wb.save(Output_file)
                wb.close() 
        # if holding == 'EUROCHEM':
        #     len_A_dict_Ech = len_A_dict
        # else:
        #     len_A_dict_SUEK = len_A_dict
        


# In[21]:


import win32com
### Отправка письма
holdind = ''
Output_file = ''
top = 0
for holding in holding_list:
    if holding == 'SUEK':
        Output_file = Output_file_SUEK
        sheet_list = ['SUEK']
        sheet_dict = {'SUEK': Output_file_SUEK}
        top = 10
    else:
        Output_file = Output_file_Ech
        sheet_list = sheets_list_Ech
        sheet_dict = sheets_dict_Ech
        top = 30

    for sheet_n in sheet_list:
        # create image
        Output_file = sheet_dict[sheet_n]
        client = win32com.client.Dispatch('Excel.Application')
        wb = client.Workbooks.Open('\\'.join([os.getcwd(),Output_file]))
        # wb = client.Workbooks.Open(Output_file)
        ws = wb.Worksheets(sheet_n)
        if sheet_n == holding:
            field = f"A1:P{top}"
        elif len_A_dict[sheet_n]<top:
            field = f"A1:F{len_A_dict[sheet_n]}"
        else:    
            field = f"A1:F{top}"
        ws.Range(field).CopyPicture(Format = 2) # screen area
        img = ImageGrab.grabclipboard()
        img.save(f'{holding}.png')
        wb.Close() # иначе табл будет открыта
        client.Quit()
        # create mail
        mailItem = olApp.CreateItem(0)
        mailItem.BodyFormat = 3
        # mail title
        mailItem.Subject = f'{sheet_n} bank limits for {report_date}' # mail head
        # mail body
        html_body = f"""<html><body><p>Dear colleagues,<br><br>
        Please find attached {holding} daily report on bank limits for {report_date}:<br><br>
        <img src="{(os.path.join(os.getcwd(), holding))}.png"><br><br>    
        Please follow the attached file for details<br><br>
        Best regards,<br>
        {signature}</p></body></html>"""
        if holding == 'SUEK':
            mailItem.To = mail_to_Suek # mail to
        else:
            to_list = [mail_to_Ech, mail_ECH_SAM, mail_ECH_NAM, mail_ECH_Europe]
            to_ech_dict = {x:y for x,y in zip(sheets_list_Ech, to_list)}
            mailItem.To = to_ech_dict[sheet_n] # mail to
            # mail attachment
        mail_attachment = Output_file
    
    
        mailItem._oleobj_.Invoke(*(64209, 0, 8, 0, olNS.Accounts.Item(mail_from)))
        mailItem.Attachments.Add(os.path.join(os.getcwd(), mail_attachment))
        mailItem.HTMLBody = html_body
        mailItem.Sensitivity  = 2
    
        # mailItem.Save()
        if Display_mail == True: ### DISPLAY
            mailItem.Display()
        if Send_mail == True: ### SEND
        
            mailItem.Send()


# In[9]:


print('The "Limits" was finished')


# Просмотр отдельных данных:

# In[10]:


# BABD_data_work.buCountryCode.unique()
# # BABD_data_work.loc[:,'buName':]


# In[11]:


# query = '''select * from "RISKACCESS"."bankAccountsBalanceDaily"
# where "buCountryCode" = 'AR'
# order by "reportDate" desc
# fetch first 3 rows only'''

# # "RISKACCESS"."bankAccountsBalanceDaily"
# # "RISKACCESS"."xxmrBankLimits"
# export_from_RISKCUSTOM(query)


# In[12]:


# BABD_data_work[BABD_data_work.buCountryCode=='AR'].to_excel('AR_bank_limits_full_data.xlsx')
# # accountStatus_list = ['active', 'mmarket']
# # BABD_data_work['balanceUsd_activmoney_market'] = 0
# # BABD_data_work.loc[BABD_data_work.accountStatus.isin(accountStatus_list), 'balanceUsd_activmoney_market'] = BABD_data_work.loc[BABD_data_work.accountStatus.isin(accountStatus_list), 'balanceUsd_mln']
# # BABD_data_work = BABD_data_work[(BABD_data_work.Active!=0)&(BABD_data_work.Limit!=0)]


