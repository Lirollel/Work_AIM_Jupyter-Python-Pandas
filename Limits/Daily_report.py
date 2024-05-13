#!/usr/bin/env python
# coding: utf-8

# In[1]:


print('The "Limits" just started running')


# In[2]:


import pandas as pd
import numpy as np
from datetime import date, timedelta
import seaborn as sns
import matplotlib.pyplot as plt
import matplotlib.axes as ax
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.drawing.image import Image
import win32com.client as win32
import os
from PIL import ImageGrab
import win32com.client

olApp = win32.Dispatch('Outlook.Application')
olNS = olApp.GetNameSpace('MAPI')

import sys
sys.path.append("C:\\Users\\KlimovaAnnaA\\Documents\\MyFiles\\Projects\\OCP")
from Defs import merge_SalesUnits
from Defs import merge_Mapping
from Defs import Period
from Defs import new_list
from Defs import export_from_RISKCUSTOM
from Defs import add_in_currency_column
from Defs import concat_columns


# In[3]:


# manual_sending = False # True/False Заполните это поле, если хотите отправить отчет даже после критичных уведомлений

# Print_qualuty_check = True # Вынести QC в отдельный excel-файл? True/False
# Display_QC_mail = True # Показать письмо QC для отправки? True/False
# Send_QC_mail = True # Создать и отправить письмо для QC? True/False

Print_to_excel = True # Создать excel-файл с расчетами? True/False
Display_mail = True # Показать письмо для отправки? True/False
Send_mail = True # Создать и отправить письмо с расчетами и графиком? True/False

mail_to = 'TarakanovMIu@aimmngt.com' # Получатель письма


# In[4]:


query = f"""
SELECT *
FROM "RISKACCESS"."bankAccountsBalanceDaily"
WHERE "reportDate" = (SELECT MAX("reportDate") FROM "RISKACCESS"."bankAccountsBalanceDaily")
"""
bankAccountsBalanceDaily_data = export_from_RISKCUSTOM(query) # export data
BABD_data_work = bankAccountsBalanceDaily_data # BABD_data_work,
BABD_data_work['balanceUsd_mln'] = BABD_data_work.balanceUsd/10**6 # balanceUsd_mln
# balanceUsd_activmoney_market
accountStatus_list = ['active', 'mmarket']
BABD_data_work['balanceUsd_activmoney_market'] = 0
BABD_data_work.loc[BABD_data_work.accountStatus.isin(accountStatus_list), 'balanceUsd_activmoney_market'] = BABD_data_work.loc[BABD_data_work.accountStatus.isin(accountStatus_list), 'balanceUsd_mln']

query = f"""
SELECT CONCAT(CONCAT("bankId", "holding"), "limitType") as "con", "activeFrom", "bankId", "limitType", "limit", "holding"
FROM "RISKACCESS"."xxmrBankLimits" 
WHERE "limitType" IN ('Transit', 'Deposit')
"""
Limit_export_data = export_from_RISKCUSTOM(query) # export data
LE_data_work = Limit_export_data.sort_values(by='activeFrom').drop_duplicates(subset='con', keep='last')  # Unique strings with lasr value for activeFrom
# pivot by bankid and mln of limit
LE_data_work['H_B_concat'] = LE_data_work.holding + '_' + LE_data_work.bankId
LE_data_group = pd.pivot_table(data=LE_data_work, index=['H_B_concat', 'bankId', 'holding'], values='limit', aggfunc='sum').reset_index()
LE_data_group.limit = LE_data_group.limit/10**6
# merge BABD_data_work and LE_data_group
BABD_data_work = BABD_data_work.reset_index(drop=True)
BABD_data_work['H_B_concat'] = BABD_data_work.holding + '_' + BABD_data_work.bankId
BABD_data_work['Limit'] = BABD_data_work.merge(LE_data_group, how='left', left_on='H_B_concat', right_on='H_B_concat', validate='many_to_one').iloc[:,-1]
# Usage and Usage_activmoney_market
BABD_data_work['Usage_activmoney_market'] = 0
BABD_data_work['Usage'] = 0
BABD_data_limit_not_na = BABD_data_work[~BABD_data_work.Limit.isna()]
BABD_data_limit_not_na['Usage_activmoney_market'] = (BABD_data_limit_not_na.balanceUsd_activmoney_market/BABD_data_limit_not_na.Limit)*100
BABD_data_limit_not_na['Usage'] = (BABD_data_limit_not_na.balanceUsd_mln/BABD_data_limit_not_na.Limit)*100
BABD_data_work[~BABD_data_work.Limit.isna()] = BABD_data_limit_not_na
BABD_data_work.loc[(BABD_data_work.Usage == np.inf) | (BABD_data_work.Usage.isna()) | (BABD_data_work.Usage == -np.inf), 'Usage'] = 0
BABD_data_work.loc[(BABD_data_work.Usage_activmoney_market == np.inf) | (BABD_data_work.Usage_activmoney_market.isna()) | (BABD_data_work.Usage_activmoney_market == -np.inf), 'Usage_activmoney_market'] = 0
BABD_data_work['Segment'] = merge_SalesUnits(df=BABD_data_work, merge_col='businessSegmentDetailed', col='buCode') # merge Segment
# Bank_name
query = """SELECT "bankId", "name", "country"
FROM "RISKACCESS"."xxmrBankLimitsBanks"
"""
data_xxmrBankLimitsBanks = export_from_RISKCUSTOM(query)
data_BLB = data_xxmrBankLimitsBanks.drop_duplicates(subset='bankId').dropna()
BABD_data_work['bank_name'] = BABD_data_work.merge(data_BLB, how='left', left_on='bankId', right_on='bankId', validate='many_to_one').iloc[:, -2]
# rounding
BABD_data_work[['Usage_activmoney_market','Usage']] = BABD_data_work[['Usage_activmoney_market','Usage']].apply(lambda x:round(x, 1))
BABD_data_work[['balanceUsd_mln','balanceUsd_activmoney_market', 'Limit']] = BABD_data_work[['balanceUsd_mln','balanceUsd_activmoney_market', 'Limit']].apply(lambda x:round(x, 2))

### TO EXCEL
BABD_data_work = BABD_data_work.rename(columns={'balanceUsd_activmoney_market':'Active', 'balanceUsd_mln':'Total', 'Usage_activmoney_market':'%_active', 'Usage':'%_total'})
# Creating Excel Writer Object from Pandas 
report_date = str(BABD_data_work.reportDate.max())[:10]
# by holding
holding_list = BABD_data_work.holding.unique().tolist()
for holding in holding_list:
    holding_data = BABD_data_work[BABD_data_work.holding == holding].reset_index(drop=True) # holding data
    # table 1
    tabel_bankName = pd.pivot_table(data=holding_data, 
                    index='bank_name', 
                    values=['Limit', 'Active', 'Total', '%_active', '%_total'], 
                    aggfunc={'Limit':'mean', 'Active':'sum', 'Total':'sum', '%_active':'sum', '%_total':'sum'},
                    fill_value=0)\
                    .reset_index()\
                    .sort_values(['Active', 'Total'], ascending=False)
    tabel_bankName = tabel_bankName[['bank_name','Limit', 'Active', '%_active', 'Total', '%_total']]
    # table 2
    table_bankCountryCode = pd.pivot_table(data=holding_data, 
                    index='bankCountryCode', 
                    values=['Active', 'Total'], 
                    aggfunc={'Active':'sum', 'Total':'sum'})\
                    .reset_index()\
                    .sort_values(['Active', 'Total'], ascending=False)
    # table 3
    table_Segment = pd.pivot_table(data=holding_data, 
                    index='Segment', 
                    values=['Active', 'Total'], 
                    aggfunc={'Active':'sum', 'Total':'sum'})\
                    .reset_index()\
                    .sort_values(['Active', 'Total'], ascending=False)
    if holding == 'SUEK':
        tabel_bankName_SUEK = tabel_bankName
        table_bankCountryCode_SUEK = table_bankCountryCode
        table_Segment_SUEK = table_Segment
    else:
        tabel_bankName_Ech = tabel_bankName
        table_bankCountryCode_Ech = table_bankCountryCode
        table_Segment_Ech = table_Segment
    # to excel
    if Print_to_excel == True:
        Output_file = '_'.join([report_date, holding,'limits_report.xlsx'])
        writer = pd.ExcelWriter(Output_file, engine='openpyxl')  
        workbook=writer.book
        pd.DataFrame({'holding':holding}, index=[1]).to_excel(writer, sheet_name=holding, index=False, header=False)
        tabel_bankName.to_excel(writer, sheet_name=holding, index=False, startrow=1)
        table_bankCountryCode.to_excel(writer, sheet_name=holding, startcol=8, index=False, startrow=1)
        table_Segment.to_excel(writer, sheet_name=holding, startcol=13, index=False, startrow=1)
        writer.close()
        if holding == 'SUEK':
            Output_file_SUEK = Output_file
        else:
            Output_file_Ech = Output_file



# In[5]:


### FORMAT
if Print_to_excel == True:
    SUEK_tables_list = [tabel_bankName_SUEK, table_bankCountryCode_SUEK, table_Segment_SUEK]
    Ech_tabels_list = [tabel_bankName_Ech, table_bankCountryCode_Ech, table_Segment_Ech]
            
    holdind = ''
    Output_file = ''
    for holding in holding_list:
        if holding == 'SUEK':
            tables_list = SUEK_tables_list
            Output_file = Output_file_SUEK
        else:
            tables_list = Ech_tabels_list
            Output_file = Output_file_Ech
        # open file
        wb = openpyxl.load_workbook(Output_file)
        ws = wb[holding]
        # color
        color_areas_list = [f"C2:C{len(tables_list[0])+2}", f"J2:J{len(tables_list[1])+2}", f"O2:O{len(tables_list[2])+2}"]
        cell_color = PatternFill(start_color='00FFCC99', end_color='00FFCC99', fill_type = "solid")
        for color_area in color_areas_list:
            for row in ws[color_area]:
                for cell in row:
                    cell.fill = cell_color
        cell_color = PatternFill(start_color='00CCFFCC', end_color='00CCFFCC', fill_type = "solid")
        for row in ws['A1:F1']:
            for cell in row:
                cell.fill = cell_color
                cell.font = Font(bold=True)
        # Borders         
        medium = Side(border_style="medium", color="000000")
        right_line_areas_list = [f"F2:F{len(tables_list[0])+2}", f"K2:K{len(tables_list[1])+2}", f"P2:P{len(tables_list[2])+2}"]
        for right_line_area in right_line_areas_list:
            for row in ws[right_line_area]:
                for cell in row:
                    cell.border = Border(top=None, left=None, right=medium, bottom=None)
        left_line_areas_list = [f"A2:A{len(tables_list[0])+2}", f"I2:I{len(tables_list[1])+2}", f"N2:N{len(tables_list[2])+2}"]
        for left_line_area in left_line_areas_list:
            for row in ws[left_line_area]:
                for cell in row:
                    cell.border = Border(top=None, left=medium, right=medium, bottom=None)
        top_line_areas_list = ['A2:F2', 'I2:K2', 'N2:P2']
        for top_line_area in top_line_areas_list:
            for row in ws[top_line_area]:
                for cell in row:
                    cell.border = Border(top=medium, left=medium, right=medium, bottom=medium)
        bottom_line_areas_list = [f'A{len(tables_list[0])+2}:F{len(tables_list[0])+2}', f'I{len(tables_list[1])+2}:K{len(tables_list[1])+2}', f'N{len(tables_list[2])+2}:P{len(tables_list[2])+2}']
        for bottom_line_area in bottom_line_areas_list:
            for row in ws[bottom_line_area]:
                for cell in row:
                    cell.border = Border(top=None, bottom=medium)
        # font color errors
        tables_list[0] = tables_list[0].reset_index(drop=True)
        # tables with deviations
        significant_deviation = tables_list[0][(((tables_list[0].Active - tables_list[0].Limit) > tables_list[0].Limit*0.1) | ((tables_list[0].Active - tables_list[0].Limit) > 10)) & ((tables_list[0].Active - tables_list[0].Limit) > 0)]
        significant_deviation_index_liist = significant_deviation.index.tolist()
        insignificant_deviation = tables_list[0][(((tables_list[0].Active - tables_list[0].Limit) <= tables_list[0].Limit*0.1) & ((tables_list[0].Active - tables_list[0].Limit) <= 10)) & ((tables_list[0].Active - tables_list[0].Limit) > 0)]
        insignificant_deviation_index_liist = insignificant_deviation.index.tolist()
        # areas lists
        significant_deviation_areas_liist = [f'A{x+3}:F{x+3}' for x in significant_deviation_index_liist] # red
        insignificant_deviation_areas_liist = [f'A{x+3}:F{x+3}' for x in insignificant_deviation_index_liist] # orange
        # color areas
        cell_color = PatternFill(start_color='00FF8080', end_color='00FF8080', fill_type = "solid")
        for color_area in significant_deviation_areas_liist:
            for row in ws[color_area]:
                for cell in row:
                    cell.fill = cell_color # red
        for color_area in insignificant_deviation_areas_liist:
            for row in ws[color_area]:
                for cell in row:
                    cell.font = Font(color="00FF9900") # orange
        # close file
        wb.save(Output_file)
        wb.close() 


# In[6]:


### Отправка письма
holdind = ''
Output_file = ''
top = 0
for holding in holding_list:
    if holding == 'SUEK':
        Output_file = Output_file_SUEK
        top = 10
    else:
        Output_file = Output_file_Ech
        top = 30
    # create image
    client = win32com.client.Dispatch("Excel.Application")
    wb = client.Workbooks.Open('\\'.join([os.getcwd(),Output_file]))
    ws = wb.Worksheets(holding)
    ws.Range(f"A1:P{top}").CopyPicture(Format = 2) # screen area
    img = ImageGrab.grabclipboard()
    img.save(f'{holding}.png')
    wb.Close() # иначе табл будет открыта
    client.Quit()
    # create mail
    mailItem = olApp.CreateItem(0)
    mailItem.BodyFormat = 3
    # mail title
    mailItem.Subject = f'{holding} bank limits for {report_date}' # mail head
    # mail body
    html_body = f"""<html><body><p>Dear colleagues,<br><br>
    Please read the attached {holding} daily report on bank limits for {report_date}:<br><br>
    <img src="{(os.path.join(os.getcwd(), holding))}.png"><br>    
    Best regards,<br>
    Maksim Tarakanov<br><br>
    Whatsapp: +7 915 161 29 12<br>
    Financial risk management</p></body></html>"""
    mailItem.To = mail_to # mail to
    mail_from = 'KlimovaAnnaA@aimmngt.com' # mail from
    # mail attachment
    mail_attachment = Output_file 


    mailItem._oleobj_.Invoke(*(64209, 0, 8, 0, olNS.Accounts.Item(mail_from)))
    mailItem.Attachments.Add(os.path.join(os.getcwd(), mail_attachment))
    mailItem.HTMLBody = html_body
    mailItem.Sensitivity  = 2

    # mailItem.Save()
    if Display_mail == True: ### DISPLAY
        mailItem.Display()
    if Send_mail == True: ### SEND
        mailItem.Send()


# In[7]:


manual_map = BABD_data_work.loc[BABD_data_work['Segment'] == 'External', ['holding', 'buCode']]
manual_map


# In[8]:


print('The "Limits" was finished')


# In[9]:


wb = openpyxl.load_workbook('2024-05-06_limits_report.xlsx')
ws = wb['EUROCHEM']


ws.autofit_column_width()

wb.save('2024-05-06_limits_report.xlsx')
wb.close()


